# -*- coding: utf-8 -*-

import os
import os.path
import time

from urllib import urlopen
import urllib2
from bs4 import BeautifulSoup


from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions


import datetime
from openpyxl import Workbook,load_workbook

import getCarDetailInfo




import robotparser, requests, json, os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
urlList = {
    'lotte' : ["https://www.lotterentacar.net/robots.txt", "https://www.lotterentacar.net/kor/longsuccession/getLongSuccessionList.do"]
}


default_save_path = '/Users/elbow/Documents/lotteRent/'
webdriver_path = '/Users/elbow/Documents/webdriver/'
pageTotal = 1
# pagingList = ["goDetail('40하 1985','mnm0313@naver.com');"]
# pagingList = ["goDetail('67하 9437','skkang@dream-plus.co.kr');"]
pagingList = [
    "goDetail('40하 1985','mnm0313@naver.com');",
    "goDetail('65하7479','min97571@naver.com');",
    "goDetail('67하4428','zamduri');",
    "goDetail('04호1112','nero1997');",
    "goDetail('64하1923','pjy6325@lycos.co.kr');",
    "goDetail('18하 8088','myworld57');",
    "goDetail('67하 9437','skkang@dream-plus.co.kr');",
    "goDetail('65하8548','0803start');",
    "goDetail('18호6849','rjckdbong87');",
    "goDetail('72호 1657','wldmsdkd1350@naver.com');",
    "goDetail('68호2953','oochen14@hanmail.net');",
    "goDetail('18하8224','shonon14@naver.com');",
    "goDetail('62호8311','wkdgmlwp1@nate.com');",
    "goDetail('37호2866','dbswlgns3533@naver.com');",
    "goDetail('18하2193','pnu88@naver.com');",
    "goDetail('44하3150','tttyuu@naver.com');",
    "goDetail('57하 8124','luv22182525@gmail.com');",
    "goDetail('64호8000','wodnd6160@naver.com');",
    "goDetail('67호7612','ljh0358');",
    "goDetail('52호8070','besswil@naver.com');",
    "goDetail('46호5775','loveisblue4');",
    "goDetail('63호 8556','RUDWNS8235');",
    "goDetail('52호8445','josk2297');",
    "goDetail('18호9453','josk2297');",
    "goDetail('58하3058','Aoouoo@naver.com');",
    "goDetail('71하9547','jhm0006');",
    "goDetail('18호1028','navy0422');",
    "goDetail('46호1016','szss');",
    "goDetail('18하2408','cktjdals0228@naver.com');",
    "goDetail('69호5418','8067439@naver.com');",
    "goDetail('44호1571','wminjung1@naver.com');",
    "goDetail('52하9384','dwight3990@naver.com');",
    "goDetail('65하 5079','anar798');",
    "goDetail('67호1588','SUNOODA');",
    "goDetail('46호3149','sexy__8022@naver.com');",
    "goDetail('46호5557','zhak369');",
    "goDetail('64하1183','JOSK2297');",
    "goDetail('68호2891','JOSK2297');",
    "goDetail('67하3567','lovelyhee7@naver.com');",
    "goDetail('52하1085','p5861498@naver.com');",
    "goDetail('64호3035','ohjimin');",
    "goDetail('57하3229','qkekwndnjs');",
    "goDetail('40하5624','hungry108');",
    "goDetail('08호2606','navy0422');",
    "goDetail('04호9244','ciw0818@naver.com');",
    "goDetail('65하8747','popo010');",
    "goDetail('52호1541','vws_mjy1@nate.com');",
    "goDetail('52호1567','aijinsong');",
    "goDetail('46하9639','woalqp');",
    "goDetail('57호5531','c713637');",
    "goDetail('64호4597','jjuuaann555');",
    "goDetail('63하6365','johyun1202@naver.com');",
    "goDetail('65호2497','sirano5b');",
    "goDetail('63하6968','gkakdlq79@naver.com');",
    "goDetail('65하2113','okhmw@naver.com');",
    "goDetail('13호2143','sud123ab');",
    "goDetail('68호2895','whtjrqls');",
    "goDetail('18호7431','wkrdms09');",
    "goDetail('18호4263','yskim6302');",
    "goDetail('35호5555','goodday0806@naver.com');",
    "goDetail('67하4777','isaiah504@naver.com');",
    "goDetail('65호9004','dhromej@naver.com');",
    "goDetail('65호6678','terlbo');",
    "goDetail('66하 5609','kyss93@naver.com');",
    "goDetail('52하5567','rishan777@naver.com');",
    "goDetail('46호1071','yoinsuk@naver.com');",
    "goDetail('69하 1991 ','emmyal');",
    "goDetail('63호1556','cpj2461');",
    "goDetail('46하3738','gunface');",
    "goDetail('65호2394','gnsdlf0112@naver.com');",
    "goDetail('04호 2878','sjlovera');",
    "goDetail('63호2845','x4545');",
]
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('headless')
chrome_options.add_argument('no-sandbox')
# ChromeDriver = webdriver.Chrome('/Users/elbow/Documents/webdriver/chromedriver', chrome_options=chrome_options)
ChromeDriver = webdriver.Chrome(webdriver_path+'chromedriver')
ChromeDriver.implicitly_wait(3)
rp = robotparser.RobotFileParser()



def canFetch(robotsUrl, loadUrl) :
    rp.set_url(robotsUrl)
    rp.read()
    possible = rp.can_fetch("*", loadUrl)
    return possible

def loadHTML(loadUrl) :
    # way1
    req = requests.get(loadUrl)
    if req.status_code == 200:
        soup = BeautifulSoup(req.text, 'html.parser')
        return soup
    else :
        raise Exception("can't load html!!!!!!!!!!!!!!!")

def getPageCount():

    # soup = loadHTML(urlList['lotte'][1])
    # last_btn = soup.select('.btn_last')[0].get('onclick')

    ChromeDriver.get(urlList['lotte'][1])
    htmlSource = BeautifulSoup(ChromeDriver.page_source, 'html.parser')
    last_btn = htmlSource.select('.btn_last')[0].get('onclick')
    start = last_btn.find("(")
    end = last_btn.find(")")
    pageTotal = int( last_btn[ int(start+1):int(end) ] )
    print 'pageTotal is : ', pageTotal


    for idx in range(pageTotal) :
        if ( idx > 0 ) :
            # script = 'goPage(' + str(idx+1) +')'
            # ChromeDriver.execute_script(script)

            script = "document.getElementById('pageIndex').value = %s" % str(idx+1)
            print script
            ChromeDriver.execute_script(script)
            ChromeDriver.execute_script("document.getElementById('frm').action = '/kor/longsuccession/getLongSuccessionList.do'")
            ChromeDriver.execute_script("document.getElementById('frm').submit()")


        htmlSource = BeautifulSoup(ChromeDriver.page_source, 'html.parser')
        listItem = htmlSource.select('.list_car dt a')
        for item in listItem :
            pagingList.append(item.get('onclick'))

    # for item in pagingList:
    #     print item


def getFolder(forderName):
    if not( os.path.isdir( default_save_path + 'carImage/' + forderName ) ):
        os.makedirs( os.path.join( default_save_path + 'carImage/' + forderName ) )
        # return 'making' + forderName
        # print 'making' + forderName
        print 'create folder'
        return False
    else :
        # return str(os.path.join( default_save_path + forderName ))+' already exsit!!'
        # print str(os.path.join( default_save_path + 'carImage/' + forderName ))+' already exsit!!'
        print 'already exsit!!'
        return True

def arrayToString(arr) :
    returnStr = ''
    if arr is None :
        return 'Null'
    else :
        for idx in range(len(arr)) :
            if ( idx is not 0 ) :
                returnStr += ', '+arr[idx]
            else :
                returnStr += arr[idx]
        return returnStr


def wait():
    print 'wait'
    # WebDriverWait(ChromeDriver, 5).until(
    #     expected_conditions.presence_of_element_located( (By.ID, target) )
    #     expected_conditions.presence_of_element_located( (By.CSS_SELECTOR, '.wkefwefjkwjef') )
    # )
    # list = ChromeDriver.find_element_by_css_selector('.thumb li')
    # html = ChromeDriver.find_element_by_tag_name('html')

# def imgLoaded(results) :
#     for image in results:
#         current_link = image.get_attribute("src")
#         r = requests.get(current_link)
#         try:
#             self.assertEqual(r.status_code, 200)
#         except AssertionError, e:
#             self.verificationErrors.append(current_link + ' delivered response code of ' + r.status_code)




def pageLoaded(selectType):
    try:
        element = WebDriverWait(ChromeDriver, 5).until(
            expected_conditions.presence_of_element_located( (By.CSS_SELECTOR, selectType) )
        )
        return True
    finally:

        return False

def carInfoSaveTest(argStr):
    ChromeDriver.get(urlList['lotte'][1])
    detail = argStr
    ChromeDriver.refresh()

    if pageLoaded('.list_car .img img') is True :
        ChromeDriver.execute_script(detail)
        if pageLoaded('.search_view_left .thumb img') is True :
            print 'carNumber', detail.split("'")[1].replace(' ','')




    # if pageLoaded('.search_view_left .thumb img') is True :
    #     print 'carNumber', detail.split("'")[1].replace(' ','')
    #     htmlSource = BeautifulSoup(ChromeDriver.page_source, 'html.parser')

        # 차량 정보 가져오기
        # carDetailInfo = getCarDetailInfo.find_detail_info(htmlSource)

        # print 'next rowCount : ', rowCount

        # ws.cell( rowCount,1,carDetailInfo['basicInfo']['title'] )
        # ws.cell( rowCount,2,carDetailInfo['basicInfo']['month_price'] )
        # ws.cell( rowCount,3,carDetailInfo['basicInfo']['remain_month'] )
        # ws.cell( rowCount,4,carDetailInfo['basicInfo']['date'] )
        #
        # ws.cell( rowCount,5,carDetailInfo['saleInfo']['saler'] )
        # ws.cell( rowCount,6,carDetailInfo['saleInfo']['email'] )
        # ws.cell( rowCount,7,carDetailInfo['saleInfo']['cellphone'] )
        # ws.cell( rowCount,8,carDetailInfo['saleInfo']['area'] )
        #
        # ws.cell( rowCount,9,carDetailInfo['carInfo']['carMaker'] )
        # ws.cell( rowCount,10,carDetailInfo['carInfo']['carName'] )
        # ws.cell( rowCount,11,carDetailInfo['carInfo']['carNumber'] )
        # ws.cell( rowCount,12,carDetailInfo['carInfo']['color'] )
        # ws.cell( rowCount,13,carDetailInfo['carInfo']['kind'] )
        # ws.cell( rowCount,14,carDetailInfo['carInfo']['fuel'] )
        # ws.cell( rowCount,15,carDetailInfo['carInfo']['distance'] )
        #
        # ws.cell( rowCount,16,carDetailInfo['contractInfo']['status'] )
        # ws.cell( rowCount,17,carDetailInfo['contractInfo']['product'] )
        # ws.cell( rowCount,18,carDetailInfo['contractInfo']['Distance'] )
        # ws.cell( rowCount,19,carDetailInfo['contractInfo']['takeOver'] )
        # ws.cell( rowCount,20,carDetailInfo['contractInfo']['gurantee'] )
        # ws.cell( rowCount,21,carDetailInfo['contractInfo']['advance'] )
        # ws.cell( rowCount,22,carDetailInfo['contractInfo']['indemnfication'] )
        # ws.cell( rowCount,23,carDetailInfo['contractInfo']['support'] )
        #
        # ws.cell( rowCount,24,arrayToString( carDetailInfo['optionInfo']['comfortable'] ))
        # ws.cell( rowCount,25,arrayToString( carDetailInfo['optionInfo']['safety'] ))
        # ws.cell( rowCount,26,arrayToString( carDetailInfo['optionInfo']['etc'] ))
        #
        # ws.cell( rowCount,27,carDetailInfo['desc'] )


        # ws.append([
        #     # 기본 정보 - 타이틀, 월 대여료, 남은 개월수, 기간
        #     carDetailInfo['basicInfo']['title'],
        #     carDetailInfo['basicInfo']['month_price'],
        #     carDetailInfo['basicInfo']['remain_month'],
        #     carDetailInfo['basicInfo']['date'],
        #     # 판매자 정보 - 판매자, 이메일, 휴대전화, 지역
        #     carDetailInfo['saleInfo']['saler'],
        #     carDetailInfo['saleInfo']['email'],
        #     carDetailInfo['saleInfo']['cellphone'],
        #     carDetailInfo['saleInfo']['area'],
        #     # 차량 정보 - 제조사, 차량명, 색상, 차종, 연료, 주행거리
        #     carDetailInfo['carInfo']['carMaker'],
        #     carDetailInfo['carInfo']['carName'],
        #     carDetailInfo['carInfo']['carNumber'],
        #     carDetailInfo['carInfo']['color'],
        #     carDetailInfo['carInfo']['kind'],
        #     carDetailInfo['carInfo']['fuel'],
        #     carDetailInfo['carInfo']['distance'],
        #     # 계약 정보 - 상태, 정비상품, 약정주행거리, 인수가, 보증금, 선납금, 면책금, 추가지원금
        #     carDetailInfo['contractInfo']['status'],
        #     carDetailInfo['contractInfo']['product'],
        #     carDetailInfo['contractInfo']['Distance'],
        #     carDetailInfo['contractInfo']['takeOver'],
        #     carDetailInfo['contractInfo']['gurantee'],
        #     carDetailInfo['contractInfo']['advance'],
        #     carDetailInfo['contractInfo']['indemnfication'],
        #     carDetailInfo['contractInfo']['support'],
        #     # 옵션 정보 - 편의장치, 안전장치, 기타
        #     arrayToString( carDetailInfo['optionInfo']['comfortable'] ),
        #     arrayToString( carDetailInfo['optionInfo']['safety'] ),
        #     arrayToString( carDetailInfo['optionInfo']['etc'] ),
        #     # 차량 설명
        #     carDetailInfo['desc']
        # ])


        # wb.save(fileName)

    # --------------------------------------------------------------------------------------------------------------

    # 차량 사진 다운로드
    # print detail.split("'")[1]
    # print detail.split("'")[3]
    #
    # 차량 사진 없을 경우
    # carNumber = detail.split("'")[1].replace(' ','')
    # if ( getFolder( carNumber ) == False ) :
    #     imgSrcList = htmlSource.select('.search_view_left .thumb img')
    #     # print len(imgSrcList)
    #     for idx in range( len(imgSrcList) ) :
    #         _url = 'https://www.lotterentacar.net' + imgSrcList[idx].get('src').encode('utf-8')
    #         req = urllib2.Request( _url )
    #         data = urllib2.urlopen(req)
    #         dr = data.read()
    #         imgName = ''
    #         if imgSrcList[idx].get('alt') == '' :
    #             imgName = str(idx) + '.jpg'
    #         else :
    #             imgName = imgSrcList[idx].get('alt')+str(idx)+'.jpg'
    #         print imgName
    #         with open(
    #                 os.path.join(default_save_path + 'carImage/' + carNumber.decode('utf-8'),  imgName), mode="wb"
    #         ) as f:
    #             f.write(dr)
    #             print("저장되었습니다.")

ws = None
rowCount = 1
if canFetch(urlList['lotte'][0], urlList['lotte'][1]) :

    # print 'can crawling'
    #getPageCount()
    # carInfoSaveTest("goDetail('63하5477','hyh0401@nate.com')")

    # 이미지 크롤링 중
    # for item in pagingList :
    #     carInfoSaveTest(item)




    # today = str( datetime.datetime.today().strftime('%Y-%m-%d') )
    # fileName = default_save_path + 'excel/lotteRent-'+ today + '.xlsx'
    # if os.path.exists( fileName ):
    #     wb = load_workbook(fileName)
    # else :
    #     wb = load_workbook(default_save_path + 'excel/basic.xlsx')
    # ws = wb.active
    # ws.title = today
    # rowCount = ws.max_row
    # print 'first rowCount : ', rowCount




    for idx in range(len(pagingList)) :
        rowCount = rowCount+1
        carInfoSaveTest(pagingList[idx] )

    ChromeDriver.quit()

else :
    # print "can't crawling"
    print loadHTML(urlList['lotte'][1])



