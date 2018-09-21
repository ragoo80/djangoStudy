# -*- coding: utf-8 -*-

# https://openpyxl.readthedocs.io/en/stable/


import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl import Workbook,load_workbook

# today = datetime.datetime.today().strftime('%Y-%m-%d')
# wb = Workbook()
# ws = wb.active  #sheet생성
# ws.title = 'test'
# ws.append([11,22,33])
# ws.append(['tting'])
# ws.cell(5,2,'testing value')

wb = load_workbook('test.xlsx')
ws = wb.active
ws.title = 'again'
print ws.max_row
_r = ws.max_row+1
ws.cell(_r,1,'a')



#Freeze top row
# ws.freeze_panes = 'A1'
# Freeze 1st column
# ws.freeze_panes = 'B1'
# Freeze top row and 1st column
# ws.freeze_panes = 'B2'
wb.save('test.xlsx')