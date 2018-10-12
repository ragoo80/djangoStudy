# -*- coding: utf-8 -*-

import os, os.path, re, datetime, time, robotparser, requests, json

import urllib, urllib2
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook,load_workbook
from PIL import Image as pil_img
import numpy as np

import getCarDetailInfo


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
urlList = {
    'lotte' : ["https://www.lotterentacar.net/robots.txt", "https://www.lotterentacar.net/kor/longsuccession/getLongSuccessionList.do"]
}

# print os.getcwd() => 현재 경로 리턴 => 나중에 default_save_path를 os.getcwd() 값으로 변경해야 함.
default_save_path = '/Users/elbow/Documents/lotteRent/'
webdriver_path = '/Users/elbow/Documents/webdriver/'
pageTotal = 1
overLapList = []
imgDownList = []
pagingList = []
today = str( datetime.datetime.today().strftime('%Y-%m-%d') )
scriptPath = default_save_path + 'userScript/'+ today +'_pagingData.npz'
timeOutPath = default_save_path + 'log/'+ today + 'TimeoutException.txt'
imgTimeOutPath = default_save_path + 'log/'+ today + 'ImgTimeoutException.txt'
adminScriptPath = default_save_path + 'userScript/'+ today + '_롯데렌터카스크립트.txt'
rp = robotparser.RobotFileParser()


def canFetch(robotsUrl, loadUrl) :
    rp.set_url(robotsUrl)
    rp.read()
    possible = rp.can_fetch("*", loadUrl)
    return possible

def loadHTML(loadUrl) :
    # way1
    req = requests.get(loadUrl)
    if req.status_code == 200:
        soup = BeautifulSoup(req.text, 'html.parser')
        return soup
    else :
        raise Exception("can't load html!!!!!!!!!!!!!!!")

# list = ChromeDriver.find_element_by_css_selector('.thumb li')
# html = ChromeDriver.find_element_by_tag_name('html')
# WebDriverWait(ChromeDriver, 10).until(
#     expected_conditions.presence_of_element_located( (By.CSS_SELECTOR, target) )
# )
# return True
def getPageCount():
    ChromeDriver.get(urlList['lotte'][1])
    htmlSource = BeautifulSoup(ChromeDriver.page_source, 'html.parser')
    last_btn = htmlSource.select('.btn_last')[0].get('onclick')
    start = last_btn.find("(")
    end = last_btn.find(")")
    pageTotal = int( last_btn[ int(start+1):int(end) ] )
    print 'pageTotal is : ', pageTotal

    for idx in range(pageTotal) :
        if ( idx > 0 ) :
            # script = 'goPage(' + str(idx+1) +')'
            # ChromeDriver.execute_script(script)
            script = "document.getElementById('pageIndex').value = %s" % str(idx+1)
            print script
            ChromeDriver.execute_script(script)
            ChromeDriver.execute_script("document.getElementById('frm').action = '/kor/longsuccession/getLongSuccessionList.do'")
            ChromeDriver.execute_script("document.getElementById('frm').submit()")

        htmlSource = BeautifulSoup(ChromeDriver.page_source, 'html.parser')
        listItem = htmlSource.select('.list_car dt a')
        # print listItem
        for item in listItem :
            pagingList.append(item.get('onclick'))
            # print pagingList

    print 'total pagingList : ', pagingList
    print len(pagingList)

    #관리자 확인용 스크립트 저장
    adminScriptList = ', \n'.join([ pagingList[idx].encode('utf-8') for idx in range(len(pagingList)) ])
    mode = "a" if os.path.exists(adminScriptPath) else "w"
    with open(os.path.join(adminScriptPath), mode) as script:
        script.write(adminScriptList)
        script.close
    # pagingList scriptPath에 저장
    np.savez(scriptPath, np.array(pagingList))

    # for item in pagingList:
    #     print item

def page_has_loaded(driver):
    page_state = driver.execute_script(
        'return document.readyState;'
    )
    return page_state

def getFolder(forderName):
    forderName = forderName.replace(' ','')
    print 'in getFolder : ', forderName
    if not( os.path.isdir( default_save_path + 'carImage/' + forderName ) ):
        # os.makedirs( os.path.join( default_save_path + 'carImage/' + forderName ) )
        # return 'making' + forderName
        # print 'making' + forderName
        print 'create folder'
        return False
    else :
        # return str(os.path.join( default_save_path + forderName ))+' already exsit!!'
        # print str(os.path.join( default_save_path + 'carImage/' + forderName ))+' already exsit!!'
        # print 'already exsit!!'
        return True

def arrayToString(arr) :
    returnStr = ''
    if arr is None or arr is 'Null':
        return 'Null'
    else :
        for idx in range(len(arr)) :
            if ( idx is not 0 ) :
                returnStr += ', '+arr[idx]
            else :
                returnStr += arr[idx]
        return returnStr

def validatorCar(carNumber):
    pattern = re.compile('[a-z|A-Z|ㄱ-]+')
    if len(carNumber.replace(' ','')) != 7 or re.search(pattern, carNumber.encode('utf-8') ) != None:
        print carNumber, ' is not validatorCar'
        return False
    return True

def overLap(carNumber):
    carNumber = carNumber.replace(' ','')
    if overLapList.count(carNumber) > 0 :
        print 'overLap carNumber is : ',carNumber
        return True
    else :
        overLapList.append(carNumber)
        return False

def downloadImgList():
    imgDriver_options = webdriver.ChromeOptions()
    imgDriver_options.add_argument('application-cache')
    imgDriver_options.add_argument('no-sandbox')
    # imgDriver_options.add_argument('--headless')
    imgDriver = webdriver.Chrome(webdriver_path+'chromedriver', chrome_options=imgDriver_options)
    imgDriver.get(urlList['lotte'][1])

    if page_has_loaded(imgDriver) == 'complete' :
        if imgDriver.execute_script("return !!window.jQuery;") == False :
            print 'there is no jquery'
            imgDriver.execute_script("scr = document.createElement('script');scr.type=\"text/javascript\";scr.src=\"https://code.jquery.com/jquery-1.12.4.min.js\";document.head.append(scr);")
            time.sleep(3)
        # 자동차 이미지 데이타 TimeoutException.txt 파일 있는지 확인 후, restart할것인지 결정
        reStartImgIndex = findRestart(imgDownList, imgTimeOutPath)
        for idx in range( reStartImgIndex, len(imgDownList) ) :
            carNumber = imgDownList[idx].split("'")[1]
            # print 'carNumber in downloadImgList : ', carNumber
            if validatorCar(carNumber) is True :
                imgDriver.execute_script("scr = document.createElement('script');scr.type=\"text/javascript\";scr.textContent = \"function goDetail(carNumber,userId){$('#carNumber').val(carNumber);$('#userId').val(userId);$('#frm').attr('action','/kor/longsuccession/getLongSuccessionDetail.do').submit();}\";document.head.append(scr);")
                imgDriver.execute_script(imgDownList[idx])
                htmlSource = BeautifulSoup(imgDriver.page_source, 'html.parser')
                if htmlSource.select('#error_wrap') == [] :
                    _carNumber = carNumber.replace(' ','')
                    if not( os.path.isdir( default_save_path + 'carImage/' + _carNumber ) ):
                        print 'there is not folder'
                        os.makedirs( os.path.join( default_save_path + 'carImage/' + _carNumber ) )
                    else :
                        print 'there is exsit forder : ', _carNumber

                    saveImage(carNumber, htmlSource)
                else :
                    print carNumber, " collecting error page info"
        imgDriver.quit()

def imgLoadCheck(url) :
    try:
        urllib2.urlopen(url)
        return True
    except urllib2.HTTPError, e:
        print e.code
        return False

def saveImage(carNumber, htmlSource):
    # print 'saveImage function'
    _carNumber = carNumber.replace(' ','')
    imgSrcList = htmlSource.select('.search_view_left .thumb img')
    imgName = ''
    # /publish/pcKor/images/common/bg_car_default.png
    for idx in range( len(imgSrcList) ) :
        _url = 'https://www.lotterentacar.net' + imgSrcList[idx].get('src').encode('utf-8').replace(' ','%20')
        try:
            req = urllib2.Request(_url)
            data = urllib2.urlopen(req)
            dr = data.read()
            if imgSrcList[idx].get('alt') == '' :
                if imgSrcList[idx].get('src') == '/publish/pcKor/images/common/bg_car_default.png' :
                    imgName = 'bg_car_default.jpg'
                else :
                    imgName = 'image' + str(idx) + '.jpg'
            else :
                # imgName = imgSrcList[idx].get('alt')+str(idx)+'.jpg'
                imgName = 'image'+str(idx)+'.jpg'

            # print imgSrcList[idx].get('alt')
            with open(
                    # '66하9815'-> carNumber.decode('utf-8')
                    # '46\ud6385532'-> carNumber.encode('utf-8')
                    os.path.join(default_save_path + 'carImage/' + _carNumber.encode('utf-8'),  imgName), "w"
            ) as f:
                f.write(dr)
                # print("저장되었습니다.")
        except urllib2.HTTPError as e:
            print carNumber, ' is :', e.code
            logFilePath = default_save_path + 'log/'+ today + 'image_error_list.txt'
            if os.path.exists(logFilePath) :
                with open(os.path.join(logFilePath), "a") as log:
                    log.write(_carNumber.encode('utf-8') + '\n')
                    log.close
            else :
                with open(os.path.join(logFilePath), "w") as log:
                    log.write(_carNumber.encode('utf-8') + '\n')
                    log.close

            #rgb 모드로 이미지 새로 저장
            # urllib.urlretrieve(_url)
            # image = pil_img.open( os.path.join(default_save_path + 'carImage/' + _carNumber.decode('utf-8'), imgName) )
            # rgb_image = image.convert('RGB')
            # r, g, b = rgb_image.getpixel((1, 1))
            # newImg = pil_img.new('RGB', image.size, (r, g, b))
            # newImg.putdata(image.getdata())
            # newImg.save(imgName)

def setErrorReport(arg):
    mode = "a" if os.path.exists(timeOutPath) else "w"
    with open(os.path.join(timeOutPath), mode) as log:
        log.write('when disconnected Time is : ' + str(datetime.datetime.now()) + '\n')
        log.write('now process carNumber is : ' + str(arg.encode('utf-8')) + '\n')
        log.close
    processClose()

def carInfoSave(argStr):

    try :
        ChromeDriver.get(urlList['lotte'][1])
    except TimeoutException as e:
        setErrorReport(argStr)


    detail = argStr
    if page_has_loaded(ChromeDriver) == 'complete' :
        carNumber = detail.split("'")[1]
        print 'carNumber', carNumber

        if ChromeDriver.execute_script("return !!window.jQuery;") == False :
            print 'there is no jquery'
            ChromeDriver.execute_script("scr = document.createElement('script');scr.type='text/javascript';scr.src='https://code.jquery.com/jquery-1.12.4.min.js';document.head.append(scr);")
            time.sleep(3)

        if validatorCar(carNumber) is True and overLap(carNumber) is False :
            try :
                ChromeDriver.execute_script(detail)
            except TimeoutException as e:
                setErrorReport(argStr)

            htmlSource = BeautifulSoup(ChromeDriver.page_source, 'html.parser')
            if htmlSource.select('#error_wrap') == [] :
                # print 'this page is normal page!!!'

                carDetailInfo = getCarDetailInfo.find_detail_info(htmlSource)
                # print carDetailInfo

                ws.cell( rowCount,1,carDetailInfo['basicInfo']['title'] )
                ws.cell( rowCount,2,carDetailInfo['basicInfo']['month_price'] )
                ws.cell( rowCount,3,carDetailInfo['basicInfo']['remain_month'] )
                ws.cell( rowCount,4,carDetailInfo['basicInfo']['date'] )

                ws.cell( rowCount,5,carDetailInfo['saleInfo']['saler'] )
                ws.cell( rowCount,6,carDetailInfo['saleInfo']['email'] )
                ws.cell( rowCount,7,carDetailInfo['saleInfo']['cellphone'] )
                ws.cell( rowCount,8,carDetailInfo['saleInfo']['area'] )

                ws.cell( rowCount,9,carDetailInfo['carInfo']['carMaker'] )
                ws.cell( rowCount,10,carDetailInfo['carInfo']['carName'] )
                ws.cell( rowCount,11,carDetailInfo['carInfo']['carNumber'] )
                ws.cell( rowCount,12,carDetailInfo['carInfo']['color'] )
                ws.cell( rowCount,13,carDetailInfo['carInfo']['kind'] )
                ws.cell( rowCount,14,carDetailInfo['carInfo']['fuel'] )
                ws.cell( rowCount,15,carDetailInfo['carInfo']['distance'] )

                ws.cell( rowCount,16,carDetailInfo['contractInfo']['status'] )
                ws.cell( rowCount,17,carDetailInfo['contractInfo']['product'] )
                ws.cell( rowCount,18,carDetailInfo['contractInfo']['Distance'] )
                ws.cell( rowCount,19,carDetailInfo['contractInfo']['takeOver'] )
                ws.cell( rowCount,20,carDetailInfo['contractInfo']['gurantee'] )
                ws.cell( rowCount,21,carDetailInfo['contractInfo']['advance'] )
                ws.cell( rowCount,22,carDetailInfo['contractInfo']['indemnfication'] )
                ws.cell( rowCount,23,carDetailInfo['contractInfo']['support'] )

                ws.cell( rowCount,24,arrayToString( carDetailInfo['optionInfo']['comfortable'] ))
                ws.cell( rowCount,25,arrayToString( carDetailInfo['optionInfo']['safety'] ))
                ws.cell( rowCount,26,arrayToString( carDetailInfo['optionInfo']['etc'] ))

                ws.cell( rowCount,27,carDetailInfo['desc'] )

                wb.save(fileName)

                if getFolder( carNumber.replace('.','') ) == False :
                    imgDownList.append(detail)
            else :
                print carNumber, " collecting error page info"

def processClose():
    appName = 'chromedriver'
    killApp = 'killall -9 ' + appName
    os.system(killApp)
# --------------------------------------------------------------------------------------------------------------

def findRestart(findList, openPath):
    if os.path.exists(openPath) :
        scr = open(openPath, 'r')
        txt = scr.readlines()
        if len(list(txt)) == 1 and list(txt)[0].encode('utf-8') == '\n' :
            return 0
        else :
            # ex)now process carNumber is : goDetail('04호 1250','stereodesk@naver.com'); => goDetail('04호 1250','stereodesk@naver.com');
            reStartCar = list(txt)[len(txt)-1].split(':')[1].replace('\n', '', 1).replace(' ', '', 1).decode('utf-8')
            if findList.count(reStartCar) > 0 :
                return findList.index(reStartCar)
    else :
        return 0


ws = None
rowCount = 1

chrome_options = webdriver.ChromeOptions()
# chrome_options.add_argument('--headless')
# chrome_options.add_argument('--dns-prefetch-disable')
prefs = {'disk-cache-size': 4096, 'profile.managed_default_content_settings.images': 2}
chrome_options.add_experimental_option('prefs', prefs)

chrome_options.add_argument('application-cache')
chrome_options.add_argument('no-sandbox')
ChromeDriver = ''

if canFetch(urlList['lotte'][0], urlList['lotte'][1]) :

    startTime =  datetime.datetime.now()
    ChromeDriver = webdriver.Chrome(webdriver_path+'chromedriver', chrome_options=chrome_options)
    ChromeDriver.set_page_load_timeout(150)

    # today +'_pagingData.npz' => pagingList 데이타 있으면 재사용
    if os.path.exists( scriptPath ):
        scriptData = np.load(scriptPath)
        pagingList = list(scriptData['arr_0'])
        # print len(pagingList) > 0
    else :
        getPageCount()

    fileName = default_save_path + 'excel/lotteRent-'+ today + '.xlsx'
    if os.path.exists( fileName ):
        wb = load_workbook(fileName)
    else :
        wb = load_workbook(default_save_path + 'excel/basic.xlsx')
        ws = wb.active
        ws.title = today
        rowCount = ws.max_row
        print 'first rowCount : ', rowCount

    reStartIndex = findRestart(pagingList, timeOutPath)
    for idx in range( reStartIndex, len(pagingList) ) :
        rowCount = rowCount + 1
        carInfoSave(pagingList[idx] )

    # 여기서 자동차 데이타 TimeoutException.txt 파일 지워야 함
    if os.path.exists(timeOutPath) : os.remove(timeOutPath)
    ChromeDriver.quit()
    processClose()

    if len(imgDownList) > 0 :
        downloadImgList()
    # print 'imgDownList is :', imgDownList
    if os.path.exists(imgTimeOutPath) : os.remove(imgTimeOutPath)

    endTime =  datetime.datetime.now()
    totalTime = endTime - startTime
    print 'totalTime is : ', totalTime


else :
    # print "can't crawling"
    print loadHTML(urlList['lotte'][1])


